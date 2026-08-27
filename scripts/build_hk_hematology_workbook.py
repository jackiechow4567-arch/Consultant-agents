#!/usr/bin/env python3
"""Build Excel workbook from HK private hospital haematology CSV sources."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
PM_REF = ROOT / "industry" / "pm-reference"
LOCATIONS_CSV = PM_REF / "hk-private-hospitals-hematology.csv"
DOCTORS_CSV = PM_REF / "hk-private-hospitals-hematologists.csv"
OUTPUT_XLSX = PM_REF / "hk-private-hospitals-hematology.xlsx"


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    return rows[0], rows[1:]


def autosize_columns(ws, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 12), max_width)


def write_sheet(ws, headers: list[str], rows: list[list[str]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)


def main() -> None:
    loc_headers, loc_rows = read_csv(LOCATIONS_CSV)
    doc_headers, doc_rows = read_csv(DOCTORS_CSV)

    wb = Workbook()
    ws_locations = wb.active
    ws_locations.title = "Locations & Contacts"
    write_sheet(ws_locations, loc_headers, loc_rows)

    ws_doctors = wb.create_sheet("Haematologists")
    write_sheet(ws_doctors, doc_headers, doc_rows)

    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
